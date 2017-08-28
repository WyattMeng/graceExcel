# -*- coding: utf-8 -*-
"""
Created on Fri Aug 25 13:52:47 2017

@author: Maddox.Meng
"""

#源、目标excel
path_src = 'C:\Users\maddox.meng\Desktop\E10000.xlsx'
path_dest= r'C:\Users\maddox.meng\Desktop\GET.xlsx'
path_tst= 'C:\Users\maddox.meng\Desktop\src_dest_sheetList.xlsx'
p = 'C:\Users\maddox.meng\Desktop\zzz.xlsx'

from openpyxl import load_workbook
import openpyxl; 
print openpyxl.__version__

#修改GET
wb = load_workbook(filename = path_dest, data_only = False, guess_types = False)
##print wb.get_active_sheet
ws = wb.get_sheet_by_name(u'应收利息')
print ws.cell(row=2,column=4).value
#wb.save(path_dest)


import xlrd
import xlwt
workbook = xlrd.open_workbook(path_dest)
sheet2 = workbook.sheet_by_name(u'应收利息')
print sheet2.name,sheet2.nrows,sheet2.ncols

##########################################
'''把pandas dataframe以openpyxl写入excel'''
#之前想按按单元格级别填数据，但在GET4.1里出
#现两个“小计”，导致定位不可靠。如果再写逻辑
#去记录每个小计出现的位置，那不如干脆取到整
#列写入GET，因为源、目标excel格式一样。现在
#要做的就是findArea of源，find目标起始位置
##########################################
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np

df = pd.read_excel(path_src, 'E10000', header=None)
df_3interest = df.iloc[11:15,2:3]
rows = dataframe_to_rows(df)

wb = load_workbook(filename = path_dest, data_only = False, guess_types = False)
ws = wb.get_sheet_by_name('interest')


#以坐标(4,,7)为起始点写入dataframe，整列写入ok没问题
#接下来要考虑如何找到起始位置(4,7)，即“应收存放金融机构利息”的行号，“2016年12月31日”列号
#另一个问题是，我怎么知道是找“应收存放金融机构利息”? “2016年12月31日”?
#基本上行号都是第2列中第一次出现的的中文字，但4.1就是特例。——所以，我在E10000findArea的时候\
#把空白值row作为area的一部分，整体粘贴到get。这样，E10000，get用同一种方式确定起始位置就行了
x=4
for r in dataframe_to_rows(df_3interest, index=False, header=False):
    ws.cell(row=x,column=7).value = r[0]
    x+=1
#wb.save(path_dest)

sheetContent = wb['interest']

'''注意'''
#openpyxl以坐标来定位，注意行号从1，而不是0开始。所以ws[1]代表第一行cells组成的list，那么
#A1就是ws[1][0]————这里的0代表list的1st元素，即第1列cell。
for x in range(1,ws.max_row):
    for y in range(0,ws.max_column):
        if ws[x][y].value == u'应收存放金融机构利息':
            print x,y
            #得到name所在的行号（index从1开始）


'''注意'''
#openpyxl识别类型如下：公式为str，纯字符为unicode，小数为float，整数为long
#空白为Nonetype，并非np.nan。判断方法为  if cell is not None          
for x in range(1,ws.max_row):
    for y in range(0,ws.max_column):
        cell = ws[x][y].value
#        if isinstance(cell, str) is True: #
#            print cell    #[=SUM(A1:G1)-Instraction!B10, =Instraction!B1, =SUM(D4:D5)-D7]
        if (isinstance(cell, str) is True and
            cell.find('!') != -1 and
            cell.find('+') == -1 and
            cell.find('-') == -1 and
            cell.find('SUM') == -1 and
            cell.find('*') == -1 and
            cell.find('/') == -1 ):
            #去Instraction找到B1、B5的真实值
            sheetname = cell.split('!')[0].lstrip('=')
            ws_temp = wb.get_sheet_by_name(sheetname)
            print cell, ws_temp[cell.split('!')[-1]].value,x,y
            #得到date所在的列号（index从0开始）

#那么起始位置就是ws[x][y]            
            

'''
1.spyder内置openpyxl当前版本2.4.1，怎么更新到最新
CMD里指定版本号更新，自动卸载之前的2.4.1：#######pip install openpyxl==2.5.0a3
###########重启spyder，解决了问题啦！！！'NoneType' object has no attribute 'group' load_workbook

2.pandas dataframe to existing excel by openpyxl
'''