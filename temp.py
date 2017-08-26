# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

yui = {'qqq':333, 'jjj':777, 'kkk':999}
#print yui['qqq']

import pandas as pd
path = 'C:\Users\mmmkk\Desktop\E10000.xlsx'
dest = 'C:\Users\mmmkk\Desktop\get.xlsx'
df = pd.read_excel(path,sheetname='E10000')
#print df
#pd.read_excel()

#x=0
#while x<999:
#    y=0
#    for item in df[x]:
#        if item == '应收存放金融机构利息':
#            print x,y
#        y+=1
#    x+=1 
print '==========================================================='
x=0
for item in df['e']:
    if item == u'应收存放金融机构利息':
        print x
    x+=1 
    
from openpyxl import load_workbook
import openpyxl; 
print openpyxl.__version__
 
#修改GET
wb = load_workbook(filename = dest, data_only = False, guess_types = False)
##print wb.get_active_sheet
ws = wb.get_sheet_by_name(u'应收利息')
x=2
y=4
cv = ws.cell(row=x,column=y).value
if cv.find('!') == -1:
    print cv
else:
    #print cv.split('!')[0].lstrip('=')
    sheetname = cv.split('!')[0].lstrip('=')
    #sheetname = "'"+sheetname+"'"
    #print sheetname
    #print '%s'%sheetname
    ws_temp = wb.get_sheet_by_name(sheetname) 
    #print cv.split('!')[-1] 
    print ws_temp[cv.split('!')[-1]].value    
#wb.save(dest)  