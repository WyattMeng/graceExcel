
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
import datetime
from formatString_M import formatString,datetime_toString


#源、目标excel
#path_src = 'C:\Users\maddox.meng\Desktop\E10000.xlsx'
#path_dest= r'C:\Users\maddox.meng\Desktop\GET.xlsx'
#keyword = u'3.应收利息'

def findArea(path_src, sheetname, keyword):
	
    #pandas读取E10000数据
    df = pd.read_excel(path_src, sheetname, header=None)
    #得到keyword‘3.应收利息’的行号x_min
    x_min=0
    for item in df.iloc[:,0]:
        if item is not np.nan and formatString(item) == keyword:
	         #print df.iloc[x,0] #x=8
	         break  #必须要break，否则x会到最后一行108
        x_min+=1

    #找 '3.应收利息' 往下的第一个带数字开头的item，以确定area的height, x_diff
    x_diff = 0
    for item in df.iloc[x_min+1:,0]:
        if item is not np.nan and item[0].isdigit() is True:
	         break
        x_diff+=1	
    #area最后一行行号x_max
    x_max = x_min + x_diff
	 #'3.应收利息'区域的df
	 #df_area = df[x_min:x_max+1]

    #'3.应收利息'下第一个科目的行号、名称，最后一个科目行号
    names = []
    xList = []
    for x in range(x_min+1, x_max+1):
	     if df.iloc[x,0] is not np.nan:
	         #print x,df.iloc[x,0]
	         names.append(df.iloc[x,0])
	         xList.append(x)
	        
    startSubName = names[0]
    x_start = xList[0]
    x_end = xList[-1]

    #创建一个dates字典，key为日期，value为列号
    dates={}
    for x in range(x_min, x_max+1):
        for y in range(0, df.shape[1]):  #注意这里不是df.shape[1]+1，否则IndexError: single positional indexer is out-of-bounds
            if isinstance(df.iloc[x,y],datetime.datetime) is True:            
                #print y,df.iloc[x,y]
	             date = datetime_toString(df.iloc[x,y])
	             dates[date] = {'y':y, 'df_area':df.iloc[x_start:x_end+1,y]}
	             #dates['df_area'] = df.iloc[x_start:x_end+1, y] 
	            
    return startSubName, x_start, x_end, dates#, df_area

#findArea('C:\Users\maddox.meng\Desktop\E10000.xlsx','E10000',u'3.应收利息')


#1.得到所需data Area的df
#def dfArea(path_src, sheetname, x_start, x_end, dates):
#    
#	 df = pd.read_excel(path_src, sheetname, header=None)
#	
#	 for key in dates:
#	 	 if key == CY:
#	 	 	 y = dates[key]
#            
#	#df_3interest = df.iloc[11:15,2:3]
#	 df_area = df.iloc[x_start:x_end+1, y]
#	 return df_area 
    



    
def startPoint(path_dest,sheetname,startSubName,date):

    wb = load_workbook(filename = path_dest, data_only = False, guess_types = False)
    ws = wb.get_sheet_by_name(sheetname)

    startPoint = []

    for x in range(1,ws.max_row):
        for y in range(0,ws.max_column):
            #if ws[x][y].value == u'应收存放金融机构利息':
            if ws[x][y].value == startSubName:   
                print x,y
                #得到name所在的行号（index从1开始）
                startPoint.append(x)

    for x in range(1,ws.max_row+1):
        for y in range(0,ws.max_column):
            cell = ws[x][y].value
    #        if isinstance(cell, str) is True: #
    #            print cell    #[=SUM(A1:G1)-Instraction!B10, =Instraction!B1, =SUM(D4:D5)-D7]
            if (isinstance(cell, str) is True and
                cell.find('!') != -1 and
                cell.find('+') == -1 and
                cell.find('-') == -1 and
                cell.find('SUM') == -1 and
                cell.find('*') == -1 and
                cell.find('/') == -1 ):
                #去Instraction找到B1、B5的真实值
                sheetname = cell.split('!')[0].lstrip('=')
                ws_temp = wb.get_sheet_by_name(sheetname)
                print cell, ws_temp[cell.split('!')[-1]].value,x,y
                
                
                #if ws_temp[cell.split('!')[-1]].value == u'2016年12月31日':
                if ws_temp[cell.split('!')[-1]].value == date:  
                #得到date所在的列号（index从0开始）
                    startPoint.append(y)
                    
    return startPoint                    
    
    
#def writeXlsx(startPoint,df_area):
#
#    x=startPoint[0]
#    y=startPoint[1]
#    for r in dataframe_to_rows(df_area, index=False, header=False):
#        ws.cell(row=x,column=y).value = r[0]
#        x+=1    
    
    
    
    
    
    
    
    
    
    
    











  